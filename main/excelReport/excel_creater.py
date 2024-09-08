from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Border, Side
from io import BytesIO


class ExcelCreater:
    def __init__(self):
        # Создание нового файла Excel и выбор активного листа
        self.wb = Workbook()
        self.ws = self.wb.active

        self.fill_headers = PatternFill(fgColor="99C295", fill_type="solid")
        self.fill_cell = PatternFill(fgColor="DEE0A8", fill_type="solid")
        self.fill_cell_less = PatternFill(fgColor="FF695E", fill_type="solid")
        self.fill_cell_more = PatternFill(fgColor="AFFF5E", fill_type="solid")
        self.align_headers = Alignment(horizontal="center", vertical="center")
        self.border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        self.column_widths = [4, 15, 50, 8, 25]

    def write_headers(self, headers):
        for row_idx, row in enumerate(headers, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = self.ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = self.fill_headers
                cell.alignment = self.align_headers
                cell.border = self.border

    def write_lines(self, data, start_row):
        for row_idx, row in enumerate(data, start=1):
            site_count = 0
            for col_idx, value in enumerate(row, start=1):
                if isinstance(value, list):
                    for sub_idx, sub_value in enumerate(value, start=1):
                        cell = self.ws.cell(row=row_idx + start_row, column=col_idx + sub_idx - 1 + site_count * 2,
                                            value=sub_value)
                        cell.fill = self.fill_cell
                        cell.border = self.border
                        self.ws.column_dimensions[get_column_letter(col_idx + sub_idx - 1 + site_count * 2)].width = 10
                        if not isinstance(value[2], str) and sub_idx != 2:
                            if value[2] < 0:
                                cell.fill = self.fill_cell_less
                            elif value[2] > 0:
                                cell.fill = self.fill_cell_more
                    site_count += 1
                else:
                    cell = self.ws.cell(row=row_idx + start_row, column=col_idx, value=value)
                    cell.fill = self.fill_cell
                    cell.border = self.border
                    self.ws.column_dimensions[get_column_letter(col_idx)].width = self.column_widths[col_idx - 1]

    def set_merge(self, headers):
        self.ws.merge_cells('A1:A3')
        self.ws.merge_cells('B1:B3')
        self.ws.merge_cells('C1:C3')
        self.ws.merge_cells('D1:D3')
        self.ws.merge_cells('E1:E3')

        for col_idx in range(6, len(headers[0]), 3):
            self.ws.merge_cells(f"{get_column_letter(col_idx)}1:{get_column_letter(col_idx + 2)}1")
            self.ws.merge_cells(f"{get_column_letter(col_idx)}3:{get_column_letter(col_idx + 2)}3")

    def create(self, headers, data):
        self.write_headers(headers)
        self.write_lines(data, 3)
        self.set_merge(headers)

    def create2(self, headers, data):
        self.column_widths = [4, 50, 20, 20, 20]
        self.wb.create_sheet(title="Sheet2")
        self.ws = self.wb.get_sheet_by_name("Sheet2")
        self.write_headers(headers)
        self.write_lines(data, 1)

    def get_file(self):
        excel_file = BytesIO()
        self.wb.save(excel_file)
        excel_file.seek(0)
        return excel_file
